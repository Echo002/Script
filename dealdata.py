# dealdata for BeeSwarm、Volcano、xyplot
'''
BeeSwarm Plot：
    title: protein/gender
    table head：control/50、tumor/50、cancer/50、control/60、tumor/60、cancer/60、control/70、tumor/70、cancer/70
    x：catagory/age
    y：protein
XY Plot：
    title：protein/catagory
    table head：age、protein
    x：age
    y：protein
Vocanlo Plot：
    title：catagory/catagory
    table head：FC、-log10(T-test)
'''

# Read *.xlsx file
import os
import math
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy import stats
from scipy.stats import pearsonr

output_fileName = 'indu.xlsx'
input_fileName = 'inputFile\indu.xlsx'

try:
    t = pd.DataFrame(pd.read_excel(input_fileName))  # header = 1 表示从第一行开始
except FileNotFoundError:
    print("File not exist！")
    exit()

# paramter set

t_head = t.columns
catagory2 = ['control', 'tumor', 'cancer']
catagory1 = ['control', 'case']
catagory = ['control']
gender = [0, 1]
gender_CN = ['男', '女']


# class set


class BeeSwarmPlot:
    def __init__(self, t):
        self.t = t
        self.t_head = t.columns

    def gen_protein_class(self, gender, count_protein):
        '''
        :param gender:gender class
        :param count_protein: what kind of protein you want
        :return: a list include control/tumor/cancer + 50/60/70
        '''
        r = []
        # 50
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] < 50) & (t['catagory'] == 'control'), self.t_head[count_protein]]))
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] < 50) & (t['catagory'] == 'tumor'), self.t_head[count_protein]]))
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] < 50) & (t['catagory'] == 'cancer'), self.t_head[count_protein]]))

        # 60
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] > 50) & (t['age'] < 70) & (t['catagory'] == 'control'), self.t_head[count_protein]]))
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] > 50) & (t['age'] < 70) & (t['catagory'] == 'tumor'), self.t_head[count_protein]]))
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] > 50) & (t['age'] < 70) & (t['catagory'] == 'cancer'), self.t_head[count_protein]]))

        # 70
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] > 70) & (t['catagory'] == 'control'), self.t_head[count_protein]]))
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] > 70) & (t['catagory'] == 'tumor'), self.t_head[count_protein]]))
        r.append(list(t.loc[(t['gender'] == gender) & (t['age'] > 70) & (t['catagory'] == 'cancer'), self.t_head[count_protein]]))
        return r

    def WriteSheet(self, ListTotal, ExcelPatch, SheetName):
        '''
        :param ListTotal: only list can received
        :param ExcelPatch: excel path
        :param SheetName: sheet name
        :return: No return, output a excel file, if not, create it.
        '''
        if not os.path.exists(ExcelPatch):
            NewFile = pd.DataFrame()
            NewFile.to_excel(ExcelPatch)
        data = pd.DataFrame(ListTotal)
        data = data.transpose()
        # Create a new sheet and write data
        excelWriter = pd.ExcelWriter(ExcelPatch, engine='openpyxl')
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        data.to_excel(excel_writer=excelWriter, sheet_name=SheetName, index=None)
        excelWriter.close()


class XYplot:
    def __init__(self, t):
        self.t = t
        self.t_head = t.columns


    def gen_protein_via_age(self, count_protein, median_age):
        # 在原有的xyplot基础上，将年龄细分（5一档, 10一档）
        r = []
        # median_age = list(t.loc[(t['age'] > median_age) & (t['age'] < median_age + 5), self.t_head[2]])
        # median_age = np.median(median_age)
        r.append(np.nanmedian(list(t.loc[(t['age'] >= median_age) & (t['age'] < median_age + 10) & (t['catagory'] == 'control'), self.t_head[count_protein]])))
        r.append(np.nanmedian(list(t.loc[(t['age'] >= median_age) & (t['age'] < median_age + 10) & (t['catagory'] == 'tumor'), self.t_head[count_protein]])))
        r.append(np.nanmedian(list(t.loc[(t['age'] >= median_age) & (t['age'] < median_age + 10) & (t['catagory'] == 'cancer'), self.t_head[count_protein]])))
        return r

    def gen_protein_class(self, count_protein, cata):
        '''
        :param count_protein: what kind of protein you want
        :param cata: one of "control, tumor and cancer"
        :return: a list include age + protein
        '''
        r = []
        r.append(list(t.loc[(t['catagory'] == catagory[cata]), 'age']))
        r.append(list(t.loc[(t['catagory'] == catagory[cata]), self.t_head[count_protein]]))
        return r

    def gen_protein_class_via_gender(self, count_protein, cata, gender):
        '''
        :param count_protein: what kind of protein you want
        :param cata: one of "control, tumor and cancer"
        :return: a list include age + protein
        '''
        r = []
        r.append(list(t.loc[(t['catagory'] == catagory[cata]) & (t['gender'] == gender), 'age']))
        r.append(list(t.loc[(t['catagory'] == catagory[cata]) & (t['gender'] == gender), self.t_head[count_protein]]))
        return r

    def WriteSheet(self, ListTotal, ExcelPatch, SheetName):
        '''
        :param ListTotal: only list can received
        :param ExcelPatch: excel path
        :param SheetName: sheet name
        :return: No return, output a excel file, if not, create it.
        '''
        if not os.path.exists(ExcelPatch):
            NewFile = pd.DataFrame()
            NewFile.to_excel(ExcelPatch)
        data = pd.DataFrame(ListTotal)
        data = data.transpose()
        # Create a new sheet and write data
        excelWriter = pd.ExcelWriter(ExcelPatch, engine='openpyxl')
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        data.to_excel(excel_writer=excelWriter, sheet_name=SheetName, index=None)
        excelWriter.close()

    def calcpearsonr(self, listA, listB):
        arrayA = np.array(listA)
        arrayB = np.array(listB)
        return pearsonr(arrayA, arrayB)


class Vocanloplot:
    def __init__(self, t):
        self.t = t
        self.t_head = t.columns

    def calculate_param(self, gender, count_protein):
        '''
        :param gender:gender class
        :return: a list include control/tumor/cancer + 50/60/70
        '''
        control_list = list(t.loc[(t['gender'] == gender) & (t['catagory'] == 'control'), self.t_head[count_protein]])
        control_mediam = np.nanmedian(control_list)

        tumor_list = list(t.loc[(t['gender'] == gender) & (t['catagory'] == 'tumor'), self.t_head[count_protein]])
        tumor_mediam = np.nanmedian(tumor_list)

        cancer_list = list(t.loc[(t['gender'] == gender) & (t['catagory'] == 'cancer'), self.t_head[count_protein]])
        cancer_mediam = np.nanmedian(cancer_list)

        FC_CT = control_mediam / tumor_mediam
        FC_CC = control_mediam / cancer_mediam

        T_test_CT = list(stats.ttest_ind(control_list, tumor_list, nan_policy='omit'))
        T_test_CC = list(stats.ttest_ind(control_list, cancer_list, nan_policy='omit'))
        T_test_CT[1] = - math.log(T_test_CT[1], 10)
        T_test_CC[1] = - math.log(T_test_CC[1], 10)

        return [FC_CT, FC_CC, T_test_CT[1], T_test_CC[1]]

    def WriteSheet(self, ListTotal, ExcelPatch, SheetName):
        '''
        :param ListTotal: only list can received
        :param ExcelPatch: excel path
        :param SheetName: sheet name
        :return: No return, output a excel file, if not, create it.
        '''
        if not os.path.exists(ExcelPatch):
            NewFile = pd.DataFrame()
            NewFile.to_excel(ExcelPatch)
        data = pd.DataFrame(ListTotal)
        data = data.transpose()
        # Create a new sheet and write data
        excelWriter = pd.ExcelWriter(ExcelPatch, engine='openpyxl')
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        data.to_excel(excel_writer=excelWriter, sheet_name=SheetName, index=None)
        excelWriter.close()

# function set


def calc_corr(a, b):
    a_avg = sum(a) / len(a)
    b_avg = sum(b) / len(b)
    # 计算分子，协方差————按照协方差公式，本来要除以n的，由于在相关系数中上下同时约去了n，于是可以不除以n
    cov_ab = sum([(x - a_avg) * (y - b_avg) for x, y in zip(a, b)])
    # 计算分母，方差乘积————方差本来也要除以n，在相关系数中上下同时约去了n，于是可以不除以n
    sq = math.sqrt(sum([(x - a_avg) ** 2 for x in a]) * sum([(x - b_avg) ** 2 for x in b]))
    corr_factor = cov_ab / sq
    return corr_factor


def deleteSheet(ExcelName, SheetName):
    try:
        wb = load_workbook(ExcelName)
        ws = wb[SheetName]
        wb.remove(ws)
        wb.save(ExcelName)
    except KeyError:
        print("sheet not exist！")
        exit()


def genBeeSwarm():
    plot = BeeSwarmPlot(t)
    for gender in range(2):
        for count_protein in range(4, len(t_head)):
            eData = plot.gen_protein_class(gender, count_protein)
            plot.WriteSheet(eData, r'BeeSwarm.xlsx', str(gender) + ' ' + t_head[count_protein])
    deleteSheet(r'BeeSwarm.xlsx', 'Sheet1')


def genXYplot():
    plot = XYplot(t)
    # catagory = ['control', 'tumor', 'cancer']
    for count_protein in range(4, len(t_head)):
        eData = []
        for class_cata in range(len(catagory)):
            eData = plot.gen_protein_class(count_protein, class_cata)
            # for gen in gender:
            #     eData = plot.gen_protein_class_via_gender(count_protein, class_cata, gen)
            #     # print(gen, catagory[class_cata], count_protein)
            #     # print(eData)
            #     # exit()
            # plot.WriteSheet(eData, output_fileName, t_head[count_protein] + gender_CN[gen])
            plot.WriteSheet(eData, output_fileName, t_head[count_protein])
    deleteSheet(output_fileName, 'Sheet1')


def genVolcanoplot():
    each_protein_list = []
    plot = Vocanloplot(t)
    for gender in range(2):
        for count_protein in range(4, len(t_head)):
            each_protein_list.append(plot.calculate_param(gender, count_protein))
    xct_plot = []
    yct_plot = []
    xcc_plot = []
    ycc_plot = []
    for i in range(4, len(t_head)):
        xct_plot.append(each_protein_list[i-4][0])
        xcc_plot.append(each_protein_list[i-4][1])
        yct_plot.append(each_protein_list[i-4][2])
        ycc_plot.append(each_protein_list[i-4][3])
    m = [xct_plot, yct_plot, xcc_plot, ycc_plot]
    xct_plot = []
    yct_plot = []
    xcc_plot = []
    ycc_plot = []
    for i in range(int((len(t_head)-4)/2), len(t_head)):
        xct_plot.append(each_protein_list[i-4][0])
        xcc_plot.append(each_protein_list[i-4][1])
        yct_plot.append(each_protein_list[i-4][2])
        ycc_plot.append(each_protein_list[i-4][3])
    fm = [xct_plot, yct_plot, xcc_plot, ycc_plot]
    plot.WriteSheet(m, r'Volcanoplot.xlsx', 'male')
    plot.WriteSheet(fm, r'Volcanoplot.xlsx', 'female')
    deleteSheet(r'Volcanoplot.xlsx', 'Sheet1')


def calcpearsonr():
    plot = XYplot(t)
    for gender in range(len(gender_CN)):
        for catagory_c in range(len(catagory1)):
            want2write = []
            for count_protein in range(4, len(t_head)):
                listAge = list(t.loc[(t['gender'] == gender_CN[gender]) & (t['catagory'] == catagory1[catagory_c]), t_head[3]])
                listFeature = list(t.loc[(t['gender'] == gender_CN[gender]) & (t['catagory'] == catagory1[catagory_c]), t_head[count_protein]])
                # print(listAge)
                # print(listFeature)
                # exit()
                want2write.append(calc_corr(listAge, listFeature))
            plot.WriteSheet(want2write, r'Pearsonr17_pos.xlsx', catagory1[catagory_c]+gender_CN[gender])
    deleteSheet(r'Pearsonr17_pos.xlsx', 'Sheet1')


# calcpearsonr()
# print(t.head(0))
# print(len(t_head))

# genXYplot()
# p = list(t.loc[(t['age'] >= 20) & (t['age'] < 30) & (t['catagory'] == 'control'), t_head[4]])
# print(np.median(p))
# print(p)
# print(len(p))

genXYplot()
